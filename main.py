import asyncio
import json
import os
import re
from concurrent.futures import ThreadPoolExecutor
from typing import Dict
from uuid import uuid4

import redis
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook

app = FastAPI(title="Processador de Excel - BotConversa")

app.mount("/img", StaticFiles(directory="templates/img"), name="images")

# Configurações
templates = Jinja2Templates(directory="templates")
executor = ThreadPoolExecutor(max_workers=3)

# Configuração Redis (adicione depois de 'executor = ...')
# REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379")
# redis_client = redis.from_url(REDIS_URL, decode_responses=True)

try:
    REDIS_URL = os.getenv("REDIS_URL", "redis://localhost:6379")
    redis_client = redis.from_url(REDIS_URL, decode_responses=True)
    redis_client.ping()  # Testa conexão
    USE_REDIS = True
    print("✓ Redis conectado")
except Exception as e:
    print(f"✗ Redis indisponível: {e}")
    print("⚠ Usando memória local (não funciona com múltiplos workers)")
    USE_REDIS = False
    jobs_status_fallback = {}

# Cria diretórios necessários
os.makedirs("uploads", exist_ok=True)
os.makedirs("output", exist_ok=True)

def get_job_status(job_id: str):
    if USE_REDIS:
        data = redis_client.get(f"job:{job_id}")
        if data:
            return json.loads(data)
        return None
    else:
        return jobs_status_fallback.get(job_id)

def set_job_status(job_id: str, status_data: dict):
    if USE_REDIS:
        redis_client.setex(f"job:{job_id}", 3600, json.dumps(status_data))
    else:
        jobs_status_fallback[job_id] = status_data

def update_job_progress(job_id: str, progress: int):
    job = get_job_status(job_id)
    if job:
        job['progresso'] = progress
        set_job_status(job_id, job)


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """Página principal com formulário de upload"""
    return templates.TemplateResponse("index.html", {"request": request})


@app.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    """
    Recebe arquivo Excel, inicia processamento em background
    """
    # Valida extensão
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            400, detail="Apenas arquivos Excel (.xlsx, .xls) são aceitos"
        )

    # Gera ID único para o job
    job_id = str(uuid4())

    # Salva arquivo temporariamente
    temp_path = f"uploads/{job_id}_{file.filename}"
    conteudo = await file.read()

    with open(temp_path, "wb") as f:
        f.write(conteudo)

    # Marca como processando
    set_job_status(job_id, {
        "status": "processing",
        "arquivo_original": file.filename,
        "progresso": 0,
    })

    # Executa processamento em background
    loop = asyncio.get_event_loop()
    loop.run_in_executor(
        executor, processar_excel_background, temp_path, job_id, file.filename
    )

    return {
        "success": True,
        "job_id": job_id,
        "message": "Arquivo enviado! Processamento iniciado.",
        "status_url": f"/status/{job_id}",
    }

@app.get("/status/{job_id}")
async def verificar_status(job_id: str):
    """Verifica status do processamento"""
    print(f"[STATUS] Verificando job: {job_id}")
    
    job = get_job_status(job_id)
    
    if not job:
        print(f"[STATUS] Job {job_id} NÃO ENCONTRADO!")
        raise HTTPException(404, detail="Job não encontrado")
    
    print(f"[STATUS] Status do job {job_id}: {job}")
    return job


@app.get("/download/{job_id}")
async def download_arquivo(job_id: str):
    """
    Faz download do arquivo processado
    """
    # ❌ REMOVE ISSO (você não está usando JSON)
    # progress_file = f"uploads/{job_id}_progress.json"

    # ✅ USA O DICIONÁRIO EM MEMÓRIA
    job = get_job_status(job_id)

    if not job:
        print(f"[DOWNLOAD] Job {job_id} não encontrado")
        raise HTTPException(404, detail="Job não encontrado")

    print(f"[DOWNLOAD] Job encontrado: {job}")

    # Verifica se o processamento foi concluído
    if job.get("status") != "completed":
        print(f"[DOWNLOAD] Job ainda não concluído. Status: {job.get('status')}")
        raise HTTPException(
            400,
            detail=f"Processamento ainda não finalizado. Status: {job.get('status')}",
        )

    # Pega o caminho do arquivo
    caminho = job.get("arquivo_saida")

    if not caminho:
        print(f"[DOWNLOAD] Caminho do arquivo não encontrado")
        raise HTTPException(500, detail="Caminho do arquivo não definido")

    print(f"[DOWNLOAD] Caminho: {caminho}")
    print(f"[DOWNLOAD] Caminho absoluto: {os.path.abspath(caminho)}")
    print(f"[DOWNLOAD] Arquivo existe? {os.path.exists(caminho)}")

    # Verifica se o arquivo existe
    if not os.path.exists(caminho):
        print(f"[DOWNLOAD] Arquivo não encontrado!")
        # Debug: lista arquivos no output
        if os.path.exists("output"):
            arquivos = os.listdir("output")
            print(f"[DOWNLOAD] Arquivos em output/: {arquivos}")
        raise HTTPException(404, detail=f"Arquivo não encontrado")

    # Verifica tamanho
    tamanho = os.path.getsize(caminho)
    print(f"[DOWNLOAD] Tamanho: {tamanho} bytes")

    if tamanho == 0:
        raise HTTPException(500, detail="Arquivo está vazio")

    # Valida o Excel
    try:
        wb_teste = load_workbook(caminho, read_only=True)
        linhas = wb_teste.active.max_row
        wb_teste.close()
        print(f"[DOWNLOAD] Arquivo validado: {linhas} linhas")
    except Exception as e:
        print(f"[DOWNLOAD] ERRO ao validar: {str(e)}")
        raise HTTPException(500, detail=f"Arquivo corrompido: {str(e)}")

    # Nome do arquivo
    nome_arquivo = job.get("nome_arquivo", "arquivo_processado.xlsx")

    print(f"[DOWNLOAD] Iniciando download de: {nome_arquivo}")

    # Retorna o arquivo
    return FileResponse(
        path=caminho,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=nome_arquivo,
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"',
            "Content-Length": str(tamanho),
        },
    )


def processar_excel_background(arquivo_entrada: str, job_id: str, nome_original: str):
    """
    Função que roda em thread separada
    Processa arquivo Excel e salva resultado
    """
    print(f"\n{'='*50}")
    print(f"INICIOU PROCESSAMENTO - Job ID: {job_id}")
    print(f"Arquivo: {arquivo_entrada}")
    print(f"{'='*50}\n")

    try:
        print(f"[{job_id}] Iniciando processamento de {nome_original}")

        # Lê o arquivo Excel com data_only=True para ignorar fórmulas e usar valores calculados
        with open(arquivo_entrada, "rb") as f:
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            linhas_originais = ws.max_row
            colunas_originais = ws.max_column

            update_job_progress(job_id, 10)

            # Lê headers
            headers = [
                str(cell.value).strip().lower() if cell.value else ""
                for cell in next(ws.iter_rows(min_row=1, max_row=1))
            ]

            print(f"[{job_id}] ({len(headers)}) Colunas encontradas: {headers}")

            idx = {h: i for i, h in enumerate(headers)}
            novo_dados = []

            # Verifica padrão de colunas
            padrao_3_colunas = set(["telefone", "nome", "etiquetas"]).issubset(
                set(headers)
            )
            padrao_4_colunas = set(
                ["primeiro nome", "sobrenome", "telefone", "etiquetas"]
            ).issubset(set(headers))

            if not padrao_3_colunas and not padrao_4_colunas:
                raise ValueError(
                    "Formato de planilha não reconhecido. Colunas necessárias não encontradas."
                )

            linhas_em_branco = 0
            celulas_com_formula_sem_valor = 0
            linhas_com_problema = []
            
            update_job_progress(job_id, 30)

            # Processa cada linha
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Pula linhas vazias
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    linhas_em_branco += 1
                    continue

                # Detecta células None que podem ser fórmulas não calculadas
                tem_none_suspeito = False
                for col_idx, cell in enumerate(row):
                    if cell is None and col_idx < len(headers):
                        col_name = headers[col_idx]
                        # Verifica se é uma coluna importante
                        if col_name in ['telefone', 'nome', 'primeiro nome', 'sobrenome', 'contato', 'celular']:
                            tem_none_suspeito = True
                            celulas_com_formula_sem_valor += 1
                            linhas_com_problema.append(row_idx)
                            break
                
                if tem_none_suspeito:
                    continue  # Pula essa linha

                primeiro_nome = ""
                sobrenome = ""
                telefone = ""
                etiquetas = "NomeConfirmado"

                # Processa nome (padrão 3 colunas: nome completo em uma coluna)
                if padrao_3_colunas:
                    nome = str(row[idx["nome"]] or "").strip()
                    partes = nome.split()
                    primeiro_nome = partes[0].title() if partes else ""
                    sobrenome = " ".join(partes[1:]).title() if len(partes) > 1 else ""

                # Processa nome (padrão 4 colunas: nome e sobrenome separados)
                elif padrao_4_colunas:
                    primeiro = str(row[idx["primeiro nome"]] or "").strip()
                    sobrenome_original = str(row[idx["sobrenome"]] or "").strip()

                    partes = primeiro.split()
                    primeiro_nome = partes[0].title() if partes else ""
                    sobrenome_splitado = (
                        " ".join(partes[1:]).title() if len(partes) > 1 else ""
                    )
                    sobrenome = (
                        f"{sobrenome_splitado} {sobrenome_original}".strip().title()
                    )

                # Processa telefone
                for col in ["telefone", "contato", "celular"]:
                    if col in idx and len(row) > idx[col]:
                        telefone = str(row[idx[col]] or "")
                        telefone = re.sub(r"\D", "", telefone)
                        while len(telefone) > 13:
                            telefone = telefone[:4] + telefone[5:]
                        break

                # Processa etiquetas
                for col in ["etiquetas", "etiqueta", "tag"]:
                    if col in idx and len(row) > idx[col]:
                        etiqueta_padrao = "NomeConfirmado"
                        val = str(row[idx[col]] or "").strip()
                        etiquetas = (
                            f"{val}, {etiqueta_padrao}"
                            if val and val.lower() != "nan"
                            else etiqueta_padrao
                        )
                        break

                # Adiciona linha se tiver telefone válido
                if telefone and len(telefone) >= 10:
                    novo_dados.append([primeiro_nome, sobrenome, telefone, etiquetas])

                # Atualiza progresso a cada 1000 linhas
                if row_idx % 1000 == 0 and linhas_originais > 0:
                    progresso = 30 + int((row_idx / linhas_originais) * 50)
                    update_job_progress(job_id, min(progresso, 80))

            # Fecha o workbook original
            wb.close()

            print(f"[{job_id}] Processadas {len(novo_dados)} linhas válidas")
            
            # Aviso sobre fórmulas
            if celulas_com_formula_sem_valor > 0:
                print(f"[{job_id}] ⚠ AVISO: {celulas_com_formula_sem_valor} células com possíveis fórmulas não calculadas foram ignoradas")
                print(f"[{job_id}] ⚠ Linhas afetadas (primeiras 10): {linhas_com_problema[:10]}")

        # Conta colunas em branco (reabre o arquivo)
        colunas_em_branco = 0
        with open(arquivo_entrada, "rb") as f:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active

            for col_idx in range(colunas_originais):
                coluna_vazia = True
                for row in ws.iter_rows(
                    min_row=2,
                    min_col=col_idx + 1,
                    max_col=col_idx + 1,
                    values_only=True,
                ):
                    if row[0] is not None and str(row[0]).strip() != "":
                        coluna_vazia = False
                        break
                if coluna_vazia:
                    colunas_em_branco += 1

            wb.close()

        update_job_progress(job_id, 85)

        # Validação antes de criar arquivo
        if len(novo_dados) == 0:
            erro_msg = "Nenhuma linha válida encontrada para processar."
            if celulas_com_formula_sem_valor > 0:
                erro_msg += f"\n\nDetectadas {celulas_com_formula_sem_valor} células com possíveis fórmulas não calculadas."
                erro_msg += "\n\nSOLUÇÃO: Abra o arquivo no Excel, pressione F9 para recalcular todas as fórmulas, salve o arquivo e tente novamente."
            raise ValueError(erro_msg)

        print(f"[{job_id}] Criando novo arquivo Excel...")

        # Cria novo workbook
        wb_novo = Workbook()
        ws_novo = wb_novo.active
        ws_novo.title = "Contatos"

        # Adiciona cabeçalho
        ws_novo.append(["Primeiro nome", "Sobrenome", "Telefone", "Etiquetas"])

        # Adiciona dados
        for linha in novo_dados:
            ws_novo.append(linha)

        # Define nome e caminho do arquivo de saída
        nome_base = os.path.splitext(nome_original)[0]
        nome_saida = f"{nome_base}_{uuid4()}.xlsx"
        caminho_saida = os.path.abspath(os.path.join("output", nome_saida))

        print(f"[{job_id}] Salvando em: {caminho_saida}")

        # Garante que o diretório existe
        os.makedirs("output", exist_ok=True)

        # Salva o arquivo
        wb_novo.save(caminho_saida)
        wb_novo.close()

        # Verifica se o arquivo foi criado
        if not os.path.exists(caminho_saida):
            raise Exception("Arquivo não foi criado no sistema de arquivos")

        tamanho = os.path.getsize(caminho_saida)
        print(f"[{job_id}] ✓ Arquivo salvo: {tamanho} bytes")

        # Valida o arquivo Excel gerado
        try:
            wb_teste = load_workbook(caminho_saida, read_only=True)
            linhas_teste = wb_teste.active.max_row
            wb_teste.close()
            print(f"[{job_id}] ✓ Arquivo validado: {linhas_teste} linhas")
        except Exception as e:
            print(f"[{job_id}] ✗ AVISO: Arquivo pode estar corrompido: {str(e)}")
            raise Exception(f"Arquivo gerado está corrompido: {str(e)}")

        update_job_progress(job_id, 100)

        # Prepara mensagem de aviso se houver fórmulas
        aviso_formulas = None
        if celulas_com_formula_sem_valor > 0:
            aviso_formulas = f"{celulas_com_formula_sem_valor} células com possíveis fórmulas não calculadas foram ignoradas. Linhas afetadas: {', '.join(map(str, linhas_com_problema[:10]))}"
            if len(linhas_com_problema) > 10:
                aviso_formulas += f" e mais {len(linhas_com_problema) - 10}..."

        # Marca como concluído
        resultado = {
            "status": "completed",
            "arquivo_saida": caminho_saida,
            "nome_arquivo": nome_saida,
            "arquivo_original": nome_original,
            "progresso": 100,
            "resultado": {
                "linhas_originais": linhas_originais,
                "colunas_originais": colunas_originais,
                "colunas_encontradas": headers,
                "linhas_novo": len(novo_dados),
                "linhas_em_branco": linhas_em_branco,
                "colunas_em_branco": colunas_em_branco,
            },
        }
        
        if aviso_formulas:
            resultado["aviso_formulas"] = aviso_formulas
        
        set_job_status(job_id, resultado)

        print(f"[{job_id}] ✓ Processamento concluído!")
        print(
            f"[{job_id}] Arquivo original: {linhas_originais} linhas x {colunas_originais} colunas"
        )
        print(f"[{job_id}] Novo arquivo: {len(novo_dados) + 1} linhas x 4 colunas")
        print(f"[{job_id}] Arquivo salvo em: {caminho_saida}")
        if aviso_formulas:
            print(f"[{job_id}] ⚠ {aviso_formulas}")

    except Exception as e:
        print(f"[{job_id}] ✗ Erro: {str(e)}")
        import traceback

        traceback.print_exc()

        set_job_status(job_id, {
            "status": "error",
            "error": str(e),
            "arquivo_original": nome_original,
            "progresso": 0,
        })

    finally:
        # Remove arquivo temporário
        if os.path.exists(arquivo_entrada):
            try:
                os.remove(arquivo_entrada)
                print(f"[{job_id}] Arquivo temporário removido")
            except Exception as e:
                print(f"[{job_id}] Erro ao remover temporário: {str(e)}")


if __name__ == "__main__":
    import uvicorn

    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)