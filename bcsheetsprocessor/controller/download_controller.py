import os
import unicodedata

from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook

from bcsheetsprocessor.config import OUTPUT_DIR
from bcsheetsprocessor.service import job_service


def download(job_id: str) -> FileResponse:
    """
    Faz download do arquivo processado
    """
    job = job_service.get_job_status(job_id)

    if not job:
        print(f"[DOWNLOAD] Job {job_id} não encontrado")
        raise HTTPException(404, detail="Job não encontrado")

    print(f"[DOWNLOAD] Job encontrado: {job}")

    # Verifica se o processamento foi concluído
    if job.get("status") != "completed":
        print(f"[DOWNLOAD] Job ainda não concluído. Status: {job.get('status')}")
        raise HTTPException(
            400,
            detail=f"Processamento ainda não finalizado. Status: {job.get('status')}",
        )

    # Pega o caminho do arquivo
    caminho = job.get("arquivo_saida")

    if not caminho:
        print(f"[DOWNLOAD] Caminho do arquivo não encontrado")
        raise HTTPException(500, detail="Caminho do arquivo não definido")

    print(f"[DOWNLOAD] Caminho: {caminho}")
    print(f"[DOWNLOAD] Caminho absoluto: {os.path.abspath(caminho)}")
    print(f"[DOWNLOAD] Arquivo existe? {os.path.exists(caminho)}")

    # Verifica se o arquivo existe
    if not os.path.exists(caminho):
        print(f"[DOWNLOAD] Arquivo não encontrado!")
        # Debug: lista arquivos no output
        if OUTPUT_DIR.exists():
            arquivos = os.listdir(OUTPUT_DIR)
            print(f"[DOWNLOAD] Arquivos em output/: {arquivos}")
        raise HTTPException(404, detail=f"Arquivo não encontrado")

    # Verifica tamanho
    tamanho = os.path.getsize(caminho)
    print(f"[DOWNLOAD] Tamanho: {tamanho} bytes")

    if tamanho == 0:
        raise HTTPException(500, detail="Arquivo está vazio")

    # Valida o Excel
    try:
        wb_teste = load_workbook(caminho, read_only=True)
        linhas = wb_teste.active.max_row
        wb_teste.close()
        print(f"[DOWNLOAD] Arquivo validado: {linhas} linhas")
    except Exception as e:
        print(f"[DOWNLOAD] ERRO ao validar: {str(e)}")
        raise HTTPException(500, detail=f"Arquivo corrompido: {str(e)}")

    # Nome do arquivo
    nome_arquivo = job.get("nome_arquivo", "arquivo_processado.xlsx")
    nome_arquivo = unicodedata.normalize('NFKD', nome_arquivo).encode('ascii', 'ignore').decode('ascii')

    print(f"[DOWNLOAD] Iniciando download de: {nome_arquivo}")

    # Retorna o arquivo
    return FileResponse(
        path=caminho,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=nome_arquivo,
    )