import os
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MAX_LINHAS_ODS = 1_000_000


@dataclass
class PlanilhaDados:
    """Dados normalizados de uma planilha (xlsx, xls ou ods) lidos em memória."""

    linhas: list[list] = field(default_factory=list)
    celulas_com_formula: set = field(default_factory=set)
    max_row: int = 0
    max_col: int = 0


def ler_planilha(caminho: str) -> PlanilhaDados:
    """Lê uma planilha e retorna dados normalizados, com dispatch por extensão."""
    ext = os.path.splitext(caminho)[1].lower()
    if ext == ".ods":
        return _ler_ods(caminho)
    if ext == ".xls":
        return _ler_xls(caminho)
    return _ler_xlsx(caminho)


def _ler_xlsx(caminho: str) -> PlanilhaDados:
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active
    linhas = [list(r) for r in ws.iter_rows(values_only=True)]
    max_row, max_col = ws.max_row, ws.max_column

    celulas_com_formula = set()
    wb_formulas = load_workbook(caminho, data_only=False)
    for linha in wb_formulas.active.iter_rows():
        for cell in linha:
            if cell.data_type == "f":
                celulas_com_formula.add(cell.coordinate)
    wb_formulas.close()
    wb.close()

    return PlanilhaDados(
        linhas=linhas,
        celulas_com_formula=celulas_com_formula,
        max_row=max_row,
        max_col=max_col,
    )


def _ler_xls(caminho: str) -> PlanilhaDados:
    import xlrd

    wb = xlrd.open_workbook(caminho)
    sh = wb.sheet_by_index(0)

    linhas = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            valor = sh.cell_value(r, c)
            if sh.cell_type(r, c) == xlrd.XL_CELL_DATE and isinstance(valor, float):
                valor = xlrd.xldate_as_datetime(valor, wb.datemode)
            row.append(valor)
        linhas.append(row)

    # xlrd 2.x não expõe células-fórmula pela API pública; valores em cache já vêm calculados
    return PlanilhaDados(
        linhas=linhas,
        celulas_com_formula=set(),
        max_row=sh.nrows,
        max_col=sh.ncols,
    )


def _ler_ods(caminho: str) -> PlanilhaDados:
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.teletype import extractText

    doc = load(caminho)
    tabelas = doc.spreadsheet.getElementsByType(Table)
    if not tabelas:
        return PlanilhaDados()
    tabela = tabelas[0]

    linhas = []
    celulas_com_formula = set()

    for row_el in tabela.getElementsByType(TableRow):
        rep_linhas = min(
            int(row_el.getAttribute("numberrowsrepeated") or 1),
            MAX_LINHAS_ODS - len(linhas),
        )

        row = []
        col = 0
        for cell in row_el.getElementsByType(TableCell):
            rep_col = int(cell.getAttribute("numbercolumnsrepeated") or 1)
            valor, formula_sem_valor = _valor_celula_ods(cell)

            if formula_sem_valor:
                celulas_com_formula.add(
                    f"{get_column_letter(col + 1)}{len(linhas) + 1}"
                )

            if valor is not None:
                while len(row) < col:
                    row.append(None)
                row.extend([valor] * rep_col)

            col += rep_col

        for _ in range(rep_linhas):
            linhas.append(row.copy())

    while linhas and _linha_vazia(linhas[-1]):
        linhas.pop()

    max_col = max((len(r) for r in linhas), default=0)
    for linha in linhas:
        while len(linha) < max_col:
            linha.append(None)

    return PlanilhaDados(
        linhas=linhas,
        celulas_com_formula=celulas_com_formula,
        max_row=len(linhas),
        max_col=max_col,
    )


def _valor_celula_ods(cell):
    """Retorna (valor, eh_formula_sem_valor) de uma célula ODF.

    Fórmula sem valor calculado (table:formula sem office:value-type) vira None
    e é registrada como suspeita — equivalente ao data_only=True do openpyxl.
    """
    from odf.teletype import extractText

    vtype = cell.getAttribute("valuetype")
    formula = cell.getAttribute("formula")
    texto = extractText(cell).strip()

    if formula and not vtype:
        return None, True

    if vtype in ("float", "currency", "percentage"):
        valor = cell.getAttribute("value")
        return (float(valor) if valor is not None else None), False
    if vtype == "boolean":
        return (cell.getAttribute("booleanvalue") == "true"), False
    if vtype in ("date", "time"):
        valor = cell.getAttribute("datevalue") or cell.getAttribute("timevalue")
        return (valor or None), False
    return (texto or None), False


def _linha_vazia(linha: list) -> bool:
    return all(c is None or str(c).strip() == "" for c in linha)
