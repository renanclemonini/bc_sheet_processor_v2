import asyncio
import os
import re
import unicodedata
from datetime import datetime, timezone
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from bcsheetsprocessor.config import OUTPUT_DIR, executor
from bcsheetsprocessor.service import job_service
from bcsheetsprocessor.service.failure_log import registrar_falha
from bcsheetsprocessor.service.sheet_reader import ler_planilha
from bcsheetsprocessor.service.telemetry_service import enviar_log_para_n8n

COLUNAS_NOME = ("nome", "nomes")
COLUNAS_PRIMEIRO_NOME = ("primeiro nome", "primeiros nomes")
COLUNAS_SOBRENOME = ("sobrenome", "sobrenomes")
COLUNAS_TELEFONE = (
    "telefone",
    "telefones",
    "contato",
    "contatos",
    "celular",
    "celulares",
)
COLUNAS_ETIQUETAS = ("etiquetas", "etiqueta", "tag", "tags")
COLUNAS_IMPORTANTES = (
    COLUNAS_NOME + COLUNAS_PRIMEIRO_NOME + COLUNAS_SOBRENOME + COLUNAS_TELEFONE
)

MIN_DIGITOS_TELEFONE = 10


def resolver_coluna(idx: dict, *variantes: str) -> int | None:
    """Retorna o índice da primeira variante encontrada (exato tem prioridade)"""
    for var in variantes:
        if var in idx:
            return idx[var]
    return None


def normalizar_nome_completo(nome: str) -> tuple[str, str]:
    """Separa nome completo em primeiro nome + sobrenome, ambos em Title Case"""
    partes = nome.split()
    primeiro_nome = partes[0].title() if partes else ""
    sobrenome = " ".join(partes[1:]).title() if len(partes) > 1 else ""
    return primeiro_nome, sobrenome


def normalizar_nome_separado(primeiro: str, sobrenome: str) -> tuple[str, str]:
    """Combina primeiro/sobrenome separados; excedentes do primeiro vão ao sobrenome"""
    partes = primeiro.split()
    primeiro_nome = partes[0].title() if partes else ""
    sobrenome_splitado = (
        " ".join(partes[1:]).title() if len(partes) > 1 else ""
    )
    sobrenome = (
        f"{sobrenome_splitado} {sobrenome}".strip().title()
    )
    return primeiro_nome, sobrenome


def normalizar_telefone(valor) -> str:
    """Remove tudo que não é dígito; remove zeros iniciais enquanto > 13 dígitos"""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    telefone = str(valor or "")
    telefone = re.sub(r"\D", "", telefone)
    while len(telefone) > 13 and telefone.startswith("0"):
        telefone = telefone[1:]
    return telefone


def normalizar_etiquetas(valor) -> str:
    """Concatena etiquetas; ignora 'nan' (mantém vazio)"""
    val = str(valor or "").strip()
    if val and val.lower() != "nan":
        return val
    return ""


def detectar_padrao(headers: list[str]) -> tuple[bool, bool]:
    """Detecta padrão 3 colunas (nome|telefone|etiquetas) e 4 colunas (1º|sobrenome|telefone|etiquetas)"""
    idx = {h: i for i, h in enumerate(headers)}
    padrao_3_colunas = (
        resolver_coluna(idx, *COLUNAS_NOME) is not None
        and resolver_coluna(idx, *COLUNAS_TELEFONE) is not None
        and resolver_coluna(idx, *COLUNAS_ETIQUETAS) is not None
    )
    padrao_4_colunas = (
        resolver_coluna(idx, *COLUNAS_PRIMEIRO_NOME) is not None
        and resolver_coluna(idx, *COLUNAS_SOBRENOME) is not None
        and resolver_coluna(idx, *COLUNAS_TELEFONE) is not None
        and resolver_coluna(idx, *COLUNAS_ETIQUETAS) is not None
    )
    return padrao_3_colunas, padrao_4_colunas


def linha_vazia(linha: list) -> bool:
    """True se todos os valores da linha são None ou string vazia"""
    return all(cell is None or str(cell).strip() == "" for cell in linha)


def processar_excel_background(
    arquivo_entrada: str, job_id: str, nome_original: str, dados_request: dict, loop
):
    """
    Função que roda em thread separada
    Processa arquivo Excel e salva resultado
    """
    print(f"\n{'='*50}")
    print(f"INICIOU PROCESSAMENTO - Job ID: {job_id}")
    print(f"Arquivo: {arquivo_entrada}")
    print(f"{'='*50}\n")

    started_at = datetime.now(timezone.utc)

    payload_telemetria = {
        "job_id": job_id,
        "status": "failed",
        "arquivo_original": nome_original,
        "arquivo_saida_nome": None,
        "tamanho_saida_bytes": None,
        "colunas_encontradas": None,
        "linhas_originais": None,
        "colunas_originais": None,
        "linhas_novo": None,
        "linhas_em_branco": None,
        "colunas_em_branco": None,
        "erro_mensagem": None,
    }

    try:
        print(f"[{job_id}] Iniciando processamento de {nome_original}")

        headers = []

        estagio = "leitura"
        dados = ler_planilha(arquivo_entrada)
        linhas = dados.linhas
        celulas_com_formula = dados.celulas_com_formula
        linhas_originais = dados.max_row
        colunas_originais = dados.max_col

        job_service.update_job_progress(job_id, 10)

        estagio = "processamento"
        headers = [
            str(valor).strip().lower() if valor else ""
            for valor in (linhas[0] if linhas else [])
        ]

        print(f"[{job_id}] ({len(headers)}) Colunas encontradas: {headers}")

        idx = {h: i for i, h in enumerate(headers)}
        novo_dados = []

        # Verifica padrão de colunas (aceita singular e plural)
        padrao_3_colunas, padrao_4_colunas = detectar_padrao(headers)

        if not padrao_3_colunas and not padrao_4_colunas:
            colunas_str = ", ".join(s for s in headers if s) or "(nenhuma coluna com nome encontrada)"
            raise ValueError(
                "Formato de planilha não reconhecido."
                f"\n\nColunas encontradas: {colunas_str}."
            )

        linhas_em_branco = 0
        celulas_com_formula_sem_valor = 0
        linhas_com_problema = []
        linhas_telefone_invalido = 0
        linhas_com_telefone_invalido = []

        col_nome = resolver_coluna(idx, *COLUNAS_NOME)
        col_primeiro_nome = resolver_coluna(idx, *COLUNAS_PRIMEIRO_NOME)
        col_sobrenome = resolver_coluna(idx, *COLUNAS_SOBRENOME)
        col_telefone = resolver_coluna(idx, *COLUNAS_TELEFONE)
        col_etiquetas = resolver_coluna(idx, *COLUNAS_ETIQUETAS)

        job_service.update_job_progress(job_id, 30)

        # Processa cada linha
        for row_idx, row in enumerate(linhas[1:], start=2):
            # Pula linhas vazias
            if linha_vazia(row):
                linhas_em_branco += 1
                continue

            # Detecta células None que podem ser fórmulas não calculadas
            tem_none_suspeito = False
            for col_idx, cell in enumerate(row):
                if cell is None and col_idx < len(headers):
                    col_name = headers[col_idx]
                    # Verifica se é uma coluna importante E contém fórmula real
                    if (
                        col_name in COLUNAS_IMPORTANTES
                        and f"{get_column_letter(col_idx + 1)}{row_idx}"
                        in celulas_com_formula
                    ):
                        tem_none_suspeito = True
                        celulas_com_formula_sem_valor += 1
                        linhas_com_problema.append(row_idx)
                        break

            if tem_none_suspeito:
                continue  # Pula essa linha

            primeiro_nome = ""
            sobrenome = ""
            telefone = ""
            etiquetas = ""

            # Processa nome (padrão 3 colunas: nome completo em uma coluna)
            if padrao_3_colunas:
                nome = str(row[col_nome] or "").strip()
                primeiro_nome, sobrenome = normalizar_nome_completo(nome)

            # Processa nome (padrão 4 colunas: nome e sobrenome separados)
            elif padrao_4_colunas:
                primeiro = str(row[col_primeiro_nome] or "").strip()
                sobrenome_original = str(row[col_sobrenome] or "").strip()
                primeiro_nome, sobrenome = normalizar_nome_separado(
                    primeiro, sobrenome_original
                )

            # Processa telefone
            if col_telefone is not None and len(row) > col_telefone:
                telefone = normalizar_telefone(row[col_telefone])

            # Processa etiquetas
            if col_etiquetas is not None and len(row) > col_etiquetas:
                etiquetas = normalizar_etiquetas(row[col_etiquetas])

            # Adiciona linha se tiver telefone válido
            if telefone and len(telefone) >= MIN_DIGITOS_TELEFONE:
                novo_dados.append([primeiro_nome, sobrenome, telefone, etiquetas])
            else:
                # Mantém o contato mesmo com telefone inválido (ex.: menos de 10 dígitos)
                linhas_telefone_invalido += 1
                if len(linhas_com_telefone_invalido) < 10:
                    linhas_com_telefone_invalido.append(row_idx)
                novo_dados.append([primeiro_nome, sobrenome, telefone, etiquetas])

            # Atualiza progresso a cada 1000 linhas
            if row_idx % 1000 == 0 and linhas_originais > 0:
                progresso = 30 + int((row_idx / linhas_originais) * 50)
                job_service.update_job_progress(job_id, min(progresso, 80))

        print(f"[{job_id}] Processadas {len(novo_dados)} linhas válidas")

        # Aviso sobre fórmulas
        if celulas_com_formula_sem_valor > 0:
            print(f"[{job_id}] ⚠ AVISO: {celulas_com_formula_sem_valor} células com possíveis fórmulas não calculadas foram ignoradas")
            print(f"[{job_id}] ⚠ Linhas afetadas (primeiras 10): {linhas_com_problema[:10]}")

        # Aviso sobre contatos mantidos sem telefone válido
        if linhas_telefone_invalido > 0:
            print(f"[{job_id}] ⚠ AVISO: {linhas_telefone_invalido} contatos mantidos com telefone inválido (< 10 dígitos)")
            print(f"[{job_id}] ⚠ Linhas afetadas (primeiras 10): {linhas_com_telefone_invalido[:10]}")

        # Conta colunas em branco a partir dos dados em memória
        colunas_em_branco = 0
        for col_idx in range(colunas_originais):
            coluna_vazia = True
            for linha in linhas[1:]:
                if (
                    len(linha) > col_idx
                    and linha[col_idx] is not None
                    and str(linha[col_idx]).strip() != ""
                ):
                    coluna_vazia = False
                    break
            if coluna_vazia:
                colunas_em_branco += 1

        job_service.update_job_progress(job_id, 85)

        # Validação antes de criar arquivo
        if len(novo_dados) == 0:
            erro_msg = "Nenhuma linha válida encontrada para processar."
            if celulas_com_formula_sem_valor > 0:
                erro_msg += f"\n\nDetectadas {celulas_com_formula_sem_valor} células com possíveis fórmulas não calculadas."
                erro_msg += "\n\nSOLUÇÃO: Abra o arquivo no Excel, pressione F9 para recalcular todas as fórmulas, salve o arquivo e tente novamente."
            raise ValueError(erro_msg)

        print(f"[{job_id}] Criando novo arquivo Excel...")

        estagio = "geracao_saida"
        wb_novo = Workbook()
        ws_novo = wb_novo.active
        ws_novo.title = "Contatos"

        # Adiciona cabeçalho
        ws_novo.append(["Primeiro nome", "Sobrenome", "Telefone", "Etiquetas"])

        # Adiciona dados
        for linha in novo_dados:
            ws_novo.append(linha)

        # Define nome e caminho do arquivo de saída
        nome_base = os.path.splitext(nome_original)[0]
        nome_base = unicodedata.normalize('NFKD', nome_base).encode('ascii', 'ignore').decode('ascii')
        nome_saida = f"{nome_base}_{uuid4()}.xlsx"
        caminho_saida = os.path.abspath(os.path.join(OUTPUT_DIR, nome_saida))

        print(f"[{job_id}] Salvando em: {caminho_saida}")

        # Garante que o diretório existe
        os.makedirs(OUTPUT_DIR, exist_ok=True)

        # Salva o arquivo
        wb_novo.save(caminho_saida)
        wb_novo.close()

        # Verifica se o arquivo foi criado
        if not os.path.exists(caminho_saida):
            raise Exception("Arquivo não foi criado no sistema de arquivos")

        tamanho = os.path.getsize(caminho_saida)
        print(f"[{job_id}] ✓ Arquivo salvo: {tamanho} bytes")

        # Valida o arquivo Excel gerado
        try:
            wb_teste = load_workbook(caminho_saida, read_only=True)
            linhas_teste = wb_teste.active.max_row
            wb_teste.close()
            print(f"[{job_id}] ✓ Arquivo validado: {linhas_teste} linhas")
        except Exception as e:
            print(f"[{job_id}] ✗ AVISO: Arquivo pode estar corrompido: {str(e)}")
            raise Exception(f"Arquivo gerado está corrompido: {str(e)}")

        job_service.update_job_progress(job_id, 100)

        # Prepara mensagem de aviso se houver fórmulas
        aviso_formulas = None
        if celulas_com_formula_sem_valor > 0:
            aviso_formulas = f"{celulas_com_formula_sem_valor} células com possíveis fórmulas não calculadas foram ignoradas. Linhas afetadas: {', '.join(map(str, linhas_com_problema[:10]))}"
            if len(linhas_com_problema) > 10:
                aviso_formulas += f" e mais {len(linhas_com_problema) - 10}..."

        # Marca como concluído
        resultado = {
            "status": "completed",
            "arquivo_saida": caminho_saida,
            "nome_arquivo": nome_saida,
            "arquivo_original": nome_original,
            "progresso": 100,
            "resultado": {
                "linhas_originais": linhas_originais,
                "colunas_originais": colunas_originais,
                "colunas_encontradas": headers,
                "linhas_novo": len(novo_dados),
                "linhas_em_branco": linhas_em_branco,
                "linhas_telefone_invalido": linhas_telefone_invalido,
                "colunas_em_branco": colunas_em_branco,
            },
        }

        if aviso_formulas:
            resultado["aviso_formulas"] = aviso_formulas

        payload_telemetria.update({
            "status": "completed",
            "arquivo_saida_nome": nome_saida,
            "tamanho_saida_bytes": tamanho,
            "colunas_encontradas": headers,
            "linhas_originais": linhas_originais,
            "colunas_originais": colunas_originais,
            "linhas_novo": len(novo_dados),
            "linhas_em_branco": linhas_em_branco,
            "linhas_telefone_invalido": linhas_telefone_invalido,
            "colunas_em_branco": colunas_em_branco,
        })

        job_service.set_job_status(job_id, resultado)

        print(f"[{job_id}] ✓ Processamento concluído!")
        print(
            f"[{job_id}] Arquivo original: {linhas_originais} linhas x {colunas_originais} colunas"
        )
        print(f"[{job_id}] Novo arquivo: {len(novo_dados) + 1} linhas x 4 colunas")
        print(f"[{job_id}] Arquivo salvo em: {caminho_saida}")
        if aviso_formulas:
            print(f"[{job_id}] ⚠ {aviso_formulas}")
        if linhas_telefone_invalido > 0:
            print(
                f"[{job_id}] ⚠ {linhas_telefone_invalido} contatos mantidos com telefone inválido (< 10 dígitos). "
                f"Linhas: {', '.join(map(str, linhas_com_telefone_invalido[:10]))}"
            )

    except Exception as e:
        print(f"[{job_id}] ✗ Erro: {str(e)}")
        import traceback

        traceback.print_exc()

        payload_telemetria["erro_mensagem"] = str(e)

        info = registrar_falha(job_id, nome_original, e, estagio, headers)

        job_service.set_job_status(job_id, {
            "status": "error",
            "error": info.mensagem_amigavel,
            "colunas_encontradas": headers,
            "arquivo_original": nome_original,
            "progresso": 0,
        })

    finally:
        payload_telemetria["duracao_ms"] = int(
            (datetime.now(timezone.utc) - started_at).total_seconds() * 1000
        )
        payload_telemetria["request"] = dados_request

        try:
            asyncio.run_coroutine_threadsafe(
                enviar_log_para_n8n(payload_telemetria), loop
            )
        except Exception as e:
            print(f"[TELEMETRY] Erro ao agendar envio para n8n: {e}")

        # Remove arquivo temporário
        if os.path.exists(arquivo_entrada):
            try:
                os.remove(arquivo_entrada)
                print(f"[{job_id}] Arquivo temporário removido")
            except Exception as e:
                print(f"[{job_id}] Erro ao remover temporário: {str(e)}")


def submit_processamento(temp_path: str, job_id: str, nome_original: str, dados_request: dict):
    loop = asyncio.get_event_loop()
    loop.run_in_executor(
        executor,
        processar_excel_background,
        temp_path,
        job_id,
        nome_original,
        dados_request,
        loop,
    )
