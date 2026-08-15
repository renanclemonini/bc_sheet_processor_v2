import asyncio

from openpyxl import load_workbook

from tests.conftest import rodar_processamento


class TestProcessamentoSucesso:
    def test_padrao3_xlsx_metricas(self, ambi_test, fixture_path):
        job_id, status = rodar_processamento(ambi_test, fixture_path("padrao3.xlsx"), loop=asyncio.new_event_loop())
        assert status["status"] == "completed"
        assert status["progresso"] == 100
        assert status["resultado"]["linhas_originais"] == 8
        assert status["resultado"]["colunas_originais"] == 3
        assert status["resultado"]["linhas_novo"] == 6
        assert status["resultado"]["linhas_em_branco"] == 1
        assert status["resultado"]["colunas_em_branco"] == 0
        assert status["resultado"]["colunas_encontradas"] == ["telefone", "nome", "etiquetas"]

    def test_padrao3_ods_metricas(self, ambi_test, fixture_path):
        _, status = rodar_processamento(ambi_test, fixture_path("padrao3.ods"), loop=asyncio.new_event_loop())
        assert status["status"] == "completed"
        assert status["resultado"]["linhas_originais"] == 8
        assert status["resultado"]["linhas_novo"] == 6

    def test_padrao3_xls_metricas(self, ambi_test, fixture_path):
        _, status = rodar_processamento(ambi_test, fixture_path("padrao3.xls"), loop=asyncio.new_event_loop())
        assert status["status"] == "completed"
        assert status["resultado"]["linhas_originais"] == 8
        assert status["resultado"]["linhas_novo"] == 6

    def test_padrao3_primeiro_nome_e_numero_metricas(self, ambi_test, fixture_path):
        _, status = rodar_processamento(
            ambi_test, fixture_path("padrao3_primeiro_nome.xlsx"), loop=asyncio.new_event_loop()
        )
        assert status["status"] == "completed"
        assert status["resultado"]["colunas_encontradas"] == ["primeiro nome", "número", "etiquetas"]
        assert status["resultado"]["linhas_originais"] == 8
        assert status["resultado"]["linhas_novo"] == 6
        assert status["resultado"]["colunas_em_branco"] == 0

        wb = load_workbook(status["arquivo_saida"], read_only=True)
        valores = list(wb.active.iter_rows(values_only=True))
        wb.close()
        assert valores[0] == ("Primeiro nome", "Sobrenome", "Telefone", "Etiquetas")
        assert valores[1] == ("Maria", "Silva", "12998123456", "cliente")

    def test_padrao4_metricas(self, ambi_test, fixture_path):
        _, status = rodar_processamento(ambi_test, fixture_path("padrao4.xlsx"), loop=asyncio.new_event_loop())
        assert status["status"] == "completed"
        assert status["resultado"]["colunas_encontradas"] == [
            "primeiro nome", "sobrenome", "telefone", "etiquetas",
        ]
        assert status["resultado"]["linhas_novo"] == 4

    def test_saida_reabre_e_tem_4_colunas(self, ambi_test, fixture_path):
        _, status = rodar_processamento(ambi_test, fixture_path("padrao3.ods"), loop=asyncio.new_event_loop())
        saida = status["arquivo_saida"]
        wb = load_workbook(saida, read_only=True)
        ws = wb.active
        assert ws.max_column == 4
        valores = list(ws.iter_rows(values_only=True))
        wb.close()
        assert valores[0] == ("Primeiro nome", "Sobrenome", "Telefone", "Etiquetas")
        assert valores[1] == ("Maria", "Silva", "12998123456", "cliente")

    def test_nome_original_ascii_na_saida(self, ambi_test, fixture_path):
        _, status = rodar_processamento(
            ambi_test, fixture_path("padrao3.xlsx"), nome_original="planilha çedilha.xlsx",
            loop=asyncio.new_event_loop(),
        )
        assert "ç" not in status["nome_arquivo"]
        assert status["nome_arquivo"].startswith("planilha cedilha_")

    def test_arquivo_temporario_removido(self, ambi_test, fixture_path):
        rodar_processamento(ambi_test, fixture_path("padrao3.xlsx"), loop=asyncio.new_event_loop())
        assert list(ambi_test["uploads_dir"].iterdir()) == []


class TestProcessamentoTelefoneInvalido:
    def test_contato_mantido_com_aviso(self, ambi_test, fixture_path):
        _, status = rodar_processamento(
            ambi_test, fixture_path("telefone_invalido.xlsx"), loop=asyncio.new_event_loop()
        )
        assert status["status"] == "completed"
        assert status["resultado"]["linhas_telefone_invalido"] == 1
        assert status["resultado"]["linhas_novo"] == 2

        wb = load_workbook(status["arquivo_saida"], read_only=True)
        valores = list(wb.active.iter_rows(values_only=True))
        wb.close()
        assert ("Curto", None, "12345678", "cliente") in valores


class TestProcessamentoErros:
    def test_formato_nao_reconhecido(self, ambi_test, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.append(["foo", "bar"])
        wb.active.append(["1", "2"])
        caminho = tmp_path / "nao_reconhecido.xlsx"
        wb.save(caminho)
        wb.close()

        _, status = rodar_processamento(ambi_test, caminho, loop=asyncio.new_event_loop())
        assert status["status"] == "error"
        assert "Formato de planilha não reconhecido" in status["error"]
        assert "foo, bar" in status["error"]

    def test_nenhuma_linha_valida(self, ambi_test, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.append(["Telefone", "Nome", "Etiquetas"])
        wb.active.append(["", "", ""])
        wb.active.append([None, None, None])
        caminho = tmp_path / "sem_valida.xlsx"
        wb.save(caminho)
        wb.close()

        _, status = rodar_processamento(ambi_test, caminho, loop=asyncio.new_event_loop())
        assert status["status"] == "error"
        assert "Nenhuma linha válida" in status["error"]

    def test_arquivo_corrompido(self, ambi_test, tmp_path):
        caminho = tmp_path / "corrompido.xlsx"
        caminho.write_bytes(b"isso nao e um xlsx")

        _, status = rodar_processamento(ambi_test, caminho, loop=asyncio.new_event_loop())
        assert status["status"] == "error"
        assert status["error"]
        assert "corrompido" in status["error"]
        assert "File is not a zip file" not in status["error"]

    def test_arquivo_corrompido_grava_log_estruturado(self, ambi_test, tmp_path, monkeypatch):
        from bcsheetsprocessor.service import failure_log

        log_dir = tmp_path / "home" / "logs"
        monkeypatch.setattr(failure_log, "LOG_DIR", log_dir)
        monkeypatch.setattr(failure_log, "LOG_PATH", log_dir / "error-logs-sheets-processor.log")

        caminho = tmp_path / "corrompido2.xlsx"
        caminho.write_bytes(b"isso nao e um xlsx")

        _, status = rodar_processamento(ambi_test, caminho, loop=asyncio.new_event_loop())
        assert status["status"] == "error"

        import json

        assert failure_log.LOG_PATH.exists()
        linha = json.loads(failure_log.LOG_PATH.read_text(encoding="utf-8"))
        assert linha["codigo"] == "arquivo_corrompido"
        assert linha["estagio"] == "leitura"
        assert linha["job_id"]


class TestAvisoFormulas:
    def test_formula_sem_valor_gerando_aviso(self, ambi_test, fixture_path):
        _, status = rodar_processamento(
            ambi_test, fixture_path("formula_sem_valor.ods"), loop=asyncio.new_event_loop()
        )
        assert status["status"] == "completed"
        assert status["aviso_formulas"] is not None
        assert "1 células com possíveis fórmulas não calculadas" in status["aviso_formulas"]
        assert "Linhas afetadas: 5" in status["aviso_formulas"]

    def test_formula_com_valor_nao_gera_aviso(self, ambi_test, fixture_path):
        _, status = rodar_processamento(
            ambi_test, fixture_path("com_formula.ods"), loop=asyncio.new_event_loop()
        )
        assert status["status"] == "completed"
        assert status.get("aviso_formulas") is None

    def test_todas_linhas_invalidas_por_formula_da_dica_f9(self, ambi_test, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Telefone", "Nome", "Etiquetas"])
        ws.append(["=1+1", "x", "y"])
        caminho = tmp_path / "so_formula.xlsx"
        wb.save(caminho)
        wb.close()

        _, status = rodar_processamento(ambi_test, caminho, loop=asyncio.new_event_loop())
        assert status["status"] == "error"
        assert "pressione F9" in status["error"]