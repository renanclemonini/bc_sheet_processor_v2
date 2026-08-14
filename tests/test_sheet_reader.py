import pytest

from bcsheetsprocessor.service.sheet_reader import PlanilhaDados, ler_planilha

FORMATOS_PADRAO3 = ["padrao3.xlsx", "padrao3.ods", "padrao3.xls"]
FORMATOS_PADRAO4 = ["padrao4.xlsx", "padrao4.ods", "padrao4.xls"]


class TestParidadeFormatos:
    @pytest.mark.parametrize("arquivo", FORMATOS_PADRAO3)
    def test_padrao3_paridade(self, fixture_path, arquivo):
        dados = ler_planilha(str(fixture_path(arquivo)))
        assert dados.max_col == 3
        assert dados.max_row == len(dados.linhas) == 8
        assert dados.linhas[0] == ["Telefone", "Nome", "Etiquetas"]
        assert dados.linhas[0][0] == "Telefone"

    @pytest.mark.parametrize("arquivo", FORMATOS_PADRAO4)
    def test_padrao4_paridade(self, fixture_path, arquivo):
        dados = ler_planilha(str(fixture_path(arquivo)))
        assert dados.max_col == 4
        assert dados.max_row == len(dados.linhas) == 5
        assert dados.linhas[0][:2] == ["Primeiro nome", "Sobrenome"]

    @pytest.mark.parametrize("arquivo", FORMATOS_PADRAO3)
    def test_telefones_numéricos_equivalentes(self, fixture_path, arquivo):
        dados = ler_planilha(str(fixture_path(arquivo)))
        telefones = [linha[0] for linha in dados.linhas[1:] if linha[0] not in (None, "")]
        assert len(telefones) == 6
        valores = [int(t) for t in telefones]
        assert valores == [12998123456, 11987654321, 31999998888, 27988887777, 12911112222, 31912345678]


class TestLeituraXlsx:
    def test_linhas_vazias_finais_removidas(self, fixture_path):
        dados = ler_planilha(str(fixture_path("linhas_vazias_finais.xlsx")))
        assert dados.max_row == 2
        assert len(dados.linhas) == 2
        assert dados.linhas[1][1] == "maria silva"

    def test_formula_xlsx_sem_valor_vira_none_e_detectada(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Telefone", "Nome", "Etiquetas"])
        ws.append([f'=CONCATENATE("{27988887777}","")', "joão de souza", "vip"])
        caminho = tmp_path / "formula.xlsx"
        wb.save(caminho)
        wb.close()

        dados = ler_planilha(str(caminho))
        assert dados.linhas[1][0] is None
        assert "A2" in dados.celulas_com_formula


class TestLeituraOds:
    def test_formula_com_valor_cacheado_ok(self, fixture_path):
        dados = ler_planilha(str(fixture_path("com_formula.ods")))
        assert "A5" not in dados.celulas_com_formula
        assert dados.linhas[4][0] is not None

    def test_formula_sem_valor_detectada(self, fixture_path):
        dados = ler_planilha(str(fixture_path("formula_sem_valor.ods")))
        assert "A5" in dados.celulas_com_formula
        assert dados.linhas[4][0] is None

    def test_linhas_alinhadas_a_max_col(self, tmp_path):
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableCell, TableRow
        from odf.text import P

        doc = OpenDocumentSpreadsheet()
        table = Table(name="Contatos")
        for valores in [
            ["Telefone", "Nome", "Etiquetas"],
            ["12998123456", "maria", "vi p"],
            ["11987654321", "joão", None],
        ]:
            row = TableRow()
            for valor in valores:
                cell = TableCell()
                if valor is not None:
                    cell.setAttribute("valuetype", "string")
                    p = P(text=str(valor))
                    cell.addElement(p)
                row.addElement(cell)
            table.addElement(row)
        doc.spreadsheet.addElement(table)
        caminho = tmp_path / "curto.ods"
        doc.save(str(caminho))

        dados = ler_planilha(str(caminho))
        assert dados.max_col == 3
        for linha in dados.linhas:
            assert len(linha) == 3

    def test_extensao_desconhecida_cai_em_xlsx(self, monkeypatch, tmp_path):
        from bcsheetsprocessor.service import sheet_reader

        chamado = {}

        def fake_ler_xlsx(caminho):
            chamado["caminho"] = caminho
            return PlanilhaDados()

        monkeypatch.setattr(sheet_reader, "_ler_xlsx", fake_ler_xlsx)
        caminho = tmp_path / "planilha.semext"
        caminho.write_bytes(b"dados")

        ler_planilha(str(caminho))
        assert chamado["caminho"] == str(caminho)

    def test_dispatch_por_extensao(self, monkeypatch, tmp_path):
        from bcsheetsprocessor.service import sheet_reader

        chamado = {}

        def make_fake(nome):
            def fake(caminho):
                chamado[nome] = str(caminho)
                return PlanilhaDados()
            return fake

        monkeypatch.setattr(sheet_reader, "_ler_ods", make_fake("ods"))
        monkeypatch.setattr(sheet_reader, "_ler_xls", make_fake("xls"))
        monkeypatch.setattr(sheet_reader, "_ler_xlsx", make_fake("xlsx"))

        for ext in (".ODS", ".XLS", ".XLSX", ".xlsm"):
            caminho = tmp_path / f"arquivo{ext}"
            caminho.write_bytes(b"x")
            ler_planilha(str(caminho))

        assert chamado == {"ods": str(tmp_path / "arquivo.ODS"),
                           "xls": str(tmp_path / "arquivo.XLS"),
                           "xlsx": str(tmp_path / "arquivo.xlsm")}


class TestEstrutura:
    def test_planilha_vazia_ods(self, tmp_path):
        from odf.opendocument import OpenDocumentSpreadsheet

        doc = OpenDocumentSpreadsheet()
        caminho = tmp_path / "vazia.ods"
        doc.save(str(caminho))
        dados = ler_planilha(str(caminho))
        assert isinstance(dados, PlanilhaDados)
        assert dados.linhas == []