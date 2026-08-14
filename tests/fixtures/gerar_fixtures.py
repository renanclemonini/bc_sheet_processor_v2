"""Gera as fixtures de planilha usadas pelos testes.

Requer LibreOffice (soffice) para as conversões .ods/.xls.
Saída: diretório fixtures/ (mesma pasta deste script).

Uso: poetry run python tests/fixtures/gerar_fixtures.py
"""

import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).resolve().parent
SOFFICE = shutil.which("soffice") or shutil.which("libreoffice")


def _salvar_xlsx(wb: Workbook, nome: str) -> Path:
    caminho = FIXTURES_DIR / nome
    wb.save(caminho)
    wb.close()
    return caminho


def _converter(origem: Path, formato: str) -> Path:
    saida = FIXTURES_DIR / (origem.stem + formato)
    subprocess.run(
        [SOFFICE, "--headless", "--convert-to", formato[1:],
         "--outdir", str(FIXTURES_DIR), str(origem)],
        check=True,
        capture_output=True,
    )
    return saida


def gerar_padrao3() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contatos"
    ws.append(["Telefone", "Nome", "Etiquetas"])
    ws.append(["12998123456", "maria silva", "cliente"])
    ws.append(["11987654321", "joão de souza", "vip, importado"])
    ws.append(["31999998888", "ana", "cadastro"])
    ws.append(["27988887777", "carlos oliveira santos", "cliente"])
    ws.append([])
    ws.append(["12911112222", "sem etiqueta", ""])
    ws.append(["31912345678", "branco meio", None])  # célula vazia no meio
    return _salvar_xlsx(wb, "padrao3.xlsx")


def gerar_padrao4() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Primeiro nome", "Sobrenome", "Telefone", "Etiquetas"])
    ws.append(["maria", "silva", "12998123456", "cliente"])
    ws.append(["joão de", "souza", "11987654321", "vip"])
    ws.append(["ana", "", "31999998888", "cadastro"])
    ws.append(["", "oliveira", "27988887777", ""])
    return _salvar_xlsx(wb, "padrao4.xlsx")


def gerar_com_formula() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Telefone", "Nome", "Etiquetas"])
    ws.append(["12998123456", "maria silva", "cliente"])
    ws.append(["11987654321", "joão de souza", "vip"])
    ws.append(["31999998888", "ana", "cadastro"])
    ws.append([f'=CONCATENATE("{27988887777}", "")', "carlos oliveira", "cliente"])
    return _salvar_xlsx(wb, "com_formula.xlsx")


def gerar_formula_sem_valor() -> Path:
    from odf.opendocument import OpenDocumentSpreadsheet, load
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    caminho_xlsx = FIXTURES_DIR / "formula_sem_valor_base.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Telefone", "Nome", "Etiquetas"])
    ws.append(["12998123456", "maria silva", "cliente"])
    ws.append(["11987654321", "joão de souza", "vip"])
    ws.append(["31999998888", "ana", "cadastro"])
    ws.append([None, "sem telefone", "cliente"])
    _salvar_xlsx(wb, caminho_xlsx.name)

    doc = OpenDocumentSpreadsheet()
    table = Table(name="Contatos")

    def texto_conteudo(valor: str) -> P:
        p = P(text=valor)
        return p

    linhas = [
        ["Telefone", "Nome", "Etiquetas"],
        ["12998123456", "maria silva", "cliente"],
        ["11987654321", "joão de souza", "vip"],
        ["31999998888", "ana", "cadastro"],
        [None, "sem telefone", "cliente"],
    ]
    for i, linha in enumerate(linhas):
        row = TableRow()
        for c, valor in enumerate(linha):
            cell = TableCell()
            if valor is None and i == 4 and c == 0:
                cell.setAttribute("formula", "of:=CONCATENATE(\"\", \"27988887777\")")
                cell.addElement(texto_conteudo(""))
            elif valor is not None:
                cell.setAttribute("valuetype", "string")
                cell.addElement(texto_conteudo(str(valor)))
            row.addElement(cell)
        table.addElement(row)

    doc.spreadsheet.addElement(table)
    caminho = FIXTURES_DIR / "formula_sem_valor.ods"
    doc.save(str(caminho))
    caminho_xlsx.unlink()
    return caminho


def gerar_telefone_invalido() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Telefone", "Nome", "Etiquetas"])
    ws.append(["12345678", "curto", "cliente"])
    ws.append(["12998123456", "valido", "cliente"])
    return _salvar_xlsx(wb, "telefone_invalido.xlsx")


def gerar_linhas_vazias_finais() -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["Telefone", "Nome", "Etiquetas"])
    ws.append(["12998123456", "maria silva", "cliente"])
    ws.append([])
    ws.append([])
    return _salvar_xlsx(wb, "linhas_vazias_finais.xlsx")


def gerar_invalido() -> Path:
    caminho = FIXTURES_DIR / "invalido.txt"
    caminho.write_text("não é uma planilha", encoding="utf-8")
    return caminho


def main() -> None:
    if not SOFFICE:
        print("soffice não encontrado — só serão geradas fixtures .xlsx", file=sys.stderr)
    gerar_padrao3()
    gerar_padrao4()
    gerar_com_formula()
    gerar_formula_sem_valor()
    gerar_telefone_invalido()
    gerar_linhas_vazias_finais()
    gerar_invalido()

    if SOFFICE:
        _converter(f := FIXTURES_DIR / "padrao3.xlsx", ".ods") or f
        _converter(FIXTURES_DIR / "padrao3.xlsx", ".xls")
        _converter(FIXTURES_DIR / "padrao4.xlsx", ".ods")
        _converter(FIXTURES_DIR / "padrao4.xlsx", ".xls")
        _converter(FIXTURES_DIR / "com_formula.xlsx", ".ods")
        FIXTURES_DIR.joinpath("com_formula.xlsx").unlink()

    print(f"Fixtures geradas em {FIXTURES_DIR}")


if __name__ == "__main__":
    main()